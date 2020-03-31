import logging

from openpyxl import load_workbook

from kubeops_api.models import Credential
from kubeops_api.models.host import Host
from kubeops_api.tasks import sync_host_info

log = logging.getLogger("host")


class HostImporter:
    def __init__(self, path, source='excel'):
        self.source = source
        self.path = path
        self.__hosts = []

    def run(self):
        if self.source == "excel":
            self._parse_excel_to_hosts()
        self._save_hosts()

    def _parse_excel_to_hosts(self):
        wb = load_workbook(self.path)
        sheet_names = wb.sheetnames
        for s_name in sheet_names:
            sh = wb[s_name]
            rows = list(sh.rows)
            for row in rows:
                if row[0].row == 1:
                    continue
                else:
                    defaults = {
                        "name": row[0].value,
                        "ip": row[1].value,
                        "port": int(row[2].value),
                        "credential": row[3].value
                    }
                    self.__hosts.append(defaults)

    def _save_hosts(self):
        counter = {
            "success": 0,
            "skip": 0,
            "fail": 0,
        }
        for host in self.__hosts:
            try:
                c = Credential.objects.get(name=host["credential"])
                host["credential"] = c
                h, created = Host.objects.get_or_create(defaults=host, name=host["name"])
                if created:
                    log.debug("import a host: {}".format(h.name))
                    counter["success"] = counter["success"] + 1
                    sync_host_info.apply_async(args=(h.id,), host_id=h.id)
                else:
                    log.debug("host {} already exists skipped".format(h.name))
                    counter["skip"] = counter["skip"] + 1
            except Exception as e:
                log.error("host {}: import error: {}".format(host["name"], e.args))
                counter["fail"] = counter["fail"] + 1
        log.info("import host result: {} success {} skip {} fail".format(counter["success"], counter["skip"],
                                                                         counter["fail"]))
