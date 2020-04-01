import os
from http import HTTPStatus

from django.conf import settings
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
from rest_framework.generics import get_object_or_404, CreateAPIView, RetrieveAPIView
from kubeoperator.settings import MEDIA_DIR
from kubeops_api.host_import import HostImporter
from kubeops_api.mixin import PageModelViewSet
from kubeops_api.models.host import Host
from kubeops_api.serializers.host import HostSerializer
from kubeops_api.models.item import Item
from kubeops_api.models.item_resource import ItemResource

__all__ = ["HostViewSet"]


class HostImportAPIView(CreateAPIView):

    def create(self, request, *args, **kwargs):
        source = request.data.get("source", None)
        for item in source:
            importer = HostImporter(path=os.path.join(MEDIA_DIR, item))
            importer.run()
        return JsonResponse(data={"success": True}, status=HTTPStatus.CREATED)


class DownloadHostImportTemplate(RetrieveAPIView):

    def retrieve(self, request, *args, **kwargs):
        def create_template_wb():
            wb = Workbook()
            s = wb.create_sheet(index=0, title="sheet1")
            s.cell(1, 1, "name")
            s.cell(1, 2, "ip")
            s.cell(1, 3, "port")
            s.cell(1, 4, "credential")
            return wb

        def file_iterator(file_name, chunk_size=512):
            with open(file_name, 'rb') as file:
                while True:
                    c = file.read(chunk_size)
                    if c:
                        yield c
                    else:
                        break

        media_dir = settings.MEDIA_DIR
        template_path = os.path.join(media_dir, "example_host.xlsx")
        if not os.path.exists(media_dir):
            os.mkdir(media_dir)
        if not os.path.exists(template_path):
            w = create_template_wb()
            w.save(template_path)
        response = HttpResponse(file_iterator(template_path))
        response["content_type"] = 'application/octet-stream'
        response['Content-Disposition'] = "attachment; filename=example_host.xlsx"
        return response


class HostViewSet(PageModelViewSet):
    queryset = Host.objects.all()
    serializer_class = HostSerializer

    def retrieve(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        host = get_object_or_404(Host, pk=pk)
        host.gather_info(retry=1)
        return super().retrieve(request, *args, **kwargs)

    def list(self, request, *args, **kwargs):
        item_name = request.query_params.get('item', None)
        if item_name:
            item = get_object_or_404(Item, name=item_name)
            resources = ItemResource.objects.filter(item_id=item.id)
            if resources:
                self.queryset = Host.objects.filter(id__in=resources.values_list("resource_id"))
        return super().list(self, request, *args, **kwargs)
